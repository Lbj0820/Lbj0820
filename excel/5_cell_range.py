from openpyxl import load_workbook
from random import *
wb = load_workbook("sample2.xlsx") 
ws = wb.active

ws.append(['번호','영어','수학']) #a, b, c
for i in range(1, 11):
    ws.append([i, randint(0,100), randint(0,100)])
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):  
        print(ws.cell(row=x, column=y).value) 

# col_B = ws['B'] #영어 column 만 가져오기
# print(col_B)
# for cell in col_B:
#     print(cell.value, end=' ')

# col_C = ws['C'] #수학 column 만 가져오기
# print(col_B)
# for cell in col_C:
#     print(cell.value, end=' ')

# row_range = ws[2:6]
# print(row_range)
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end = ' ')
#     print()

from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] #2번쨰 줄부터 마지막 줄까지
# for rows in row_range:
#     for  cell in rows:
#       #print(cell.value)
#       print(cell.coordinate, end= ' ')
#         #print(cell.coordinate, end=" ") # A/10
#       xy= coordinate_from_string(cell.coordinate)   
#       print(xy,end=' ')
#       print(xy[0], end=' ') #A
#       print(xy[1], end=' ') #2
#     print()

#전체 rows 
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[0].value)

#전체 columns
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)
#     print()

for row in ws.iter_rows(min_row=2, max_row=11, min_col=2,max_col=3):
    print(row[0].value, row[1].value)
    print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=2,max_col=3):
    print(col)
    print(col[0].value, col[1].value)

wb.save("sample2.xlsx")