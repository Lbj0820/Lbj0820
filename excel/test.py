from openpyxl import load_workbook
wb = load_workbook('sample4.xlsx')
ws = wb.active
#엑셀에 들어갈 값들 이중 배열로 만들기
table = [['학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트'],[1,3,8,5,14,26,12],
[2,2,3,7,15,24,18],
[3,1,5,8,8,12,4],
[4,0,8,7,17,21,18],
[5,4,8,7,16,25,15],
[6,0,5,8,8,17,0],
[7,0,9,10,16,27,18],
[8,4,6,6,15,19,17],
[9,0,10,9,19,30,19],
[10,0,8,8,20,25,20]]

#table의 각 배열에 0을 두번 추가하여 리스트를 2칸늘리기
for i in range(len(table)):
    table[i].append(0)
    table[i].append(0)
    if i > 0:
        table[i][3] = 10    #table[i][3]의 값들을 10으로 설정
        table[i][7] = sum(table[i]) - table[i][0] - 2*table[i][1] + 10  # table[i][7]의 값을 구함
        #ws['H'+str(i)].value = "=SUM(C"+str(i)+ ":G" + str(i)+")+10-B"+str(i)
        #각 총점별 학점을 구함
        #단, table[i][1] > 3 면 총점에 관계없이 table[i][8]의 자리에 F값을 넣음
        if table[i][1] > 3:
            table[i][8] = 'F'
        elif table[i][7] > 90:
            table[i][8] = 'A'
        elif table[i][7] > 80:
            table[i][8] = 'B'
        elif table[i][7] > 70:
            table[i][8] = 'C'
        else:
            table[i][8] = 'D'
#각 리스트 요소를 조건에 맞게 수정
table[0][1] = '결석'
table[0][7] = '총점'
table[0][8] = '학점'

#table의 값들을 sample4.xlsx의 셀 값에 넣음
for row in range(len(table)):
    for col in range(len(table[0])):
         ws.cell(row+1, col+1, table[row][col])

wb.save('sample4.xlsx')
wb.close()