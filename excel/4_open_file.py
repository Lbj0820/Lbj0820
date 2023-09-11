from openpyxl import load_workbook
wb = load_workbook("sample2.xlsx") 
ws = wb.active


# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        value = ws.cell(row=x, column=y).value
        print("{:5}".format(value), end=" ")
    print()