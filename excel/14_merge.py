from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2부터 
ws["B2"].value = "Marged Cell"

wb.save("sample_merge.xlsx")