from openpyxl import Workbook
wb = Workbook()     #새 워크북 생성

ws = wb.create_sheet()  #ws에 시트를 만듬
ws.title = "YoutSheet"     #YourSheet로 시트 탭 제목을 정함
ws.sheet_properties.tabColor = "66ff99"     #ws시트 탭의 색을 해당 색으로 지정

ws1 = wb.create_sheet("Hersheet")       #ws1에 hersheet라는 시트를 만듬
ws1.sheet_properties.tabColor = "66ff99"    #ws1에 탭 색상을 해당 색으로 지정
ws2 = wb.create_sheet("HisSheet", 3)    #ws2에 HisSheet라는 시트를 만듬
ws2.sheet_properties.tabColor = "66ff99"

new_ws = wb["HisSheet"]

print(wb.sheetnames)
ws1['A2'] = "SW"
new_ws["A1"] = "Test"

target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
