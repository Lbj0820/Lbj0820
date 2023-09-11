# 문제) 
# 여러분은 여주대학교의 소프트웨어융합과 교수입니다.
# 여러분이 가르치는 과목의 점수 비중은 다음과 같습니다.

# - 출석 10
# - 퀴즈1 10
# - 퀴즈2 10
# - 중간고사 20
# - 기말고사 30
# - 프로젝트 20
# ------------
# - 총합 100

# 마지막 수업을 모두 마치고 이번 학기 학생들의 최종 성적을 검토하는 과정에서
# 퀴즈2 문제에 오류를 발견하여 모두 만점 처리를 하기로 하였습니다.
# 현재까지 작성된 최종 성적 데이터를 기준으로 아래와 같이 수정하시오.

# 0. 결석 횟수를 출석 점수로 수정 (결석1회당 1점 감점)
# 1. 퀴즈2 점수를 10 으로 수정    
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
# 3. 결석이 4회 이상인 학생은 총점 상관없이 F

# ※ 최종 파일명 : scores자기이름.py    scores자기이름.xlsx

# [현재까지 작성된 최종 성적 데이터]
# 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# 1,3,8,5,14,26,12
# 2,2,3,7,15,24,18
# 3,1,5,8,8,12,4
# 4,0,8,7,17,21,18
# 5,4,8,7,16,25,15
# 6,0,5,8,8,17,0
# 7,0,9,10,16,27,18
# 8,4,6,6,15,19,17
# 9,0,10,9,19,30,19
# 10,0,8,8,20,25,20

from openpyxl import load_workbook
wb = load_workbook("sample3.xlsx", data_only=True)
ws = wb.active
table = [list(row) for row in ws.values]

total_score = 0
for row in range(2,12):
    ws['D'+str(row)] = 10
    #ws['H'+str(row)].value = "=SUM(C"+str(row)+ ":G" + str(row)+")+10-B"+str(row)
    
for i in range(1,11):
    for j in range(3,7):
        total_score=sum(table[i][2:7])-table[i][1]+10
        ws['H'+str(i+1)] = total_score
        table[i][7] = total_score
        if table[i][1] >3:
            ws['I'+str(i+1)] = "F"
        elif table[i][7] >90:
            ws['I'+str(i+1)] = "A"
        elif table[i][7] >80:
            ws['I'+str(i+1)] = "B"
        elif table[i][7] >70:
            ws['I'+str(i+1)] = "C"
        else:
            ws['I'+str(i+1)] = "D"

wb.save("sample3.xlsx")
wb.close()

