from openpyxl import Workbook
wb = Workbook() 
ws = wb.active
wb.title = "MySheet"

ws["A1"] = 2
ws["A2"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value)

print(ws.cell(column=1, row=1).value)
print(ws.cell(column=2, row=1).value)

c = ws.cell(column=10, row=1).value=10
print(c)

from random import *

index = 1
for x in range(1,11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0,10000))

wb.save("smaple2.xlsx")
wb.close()