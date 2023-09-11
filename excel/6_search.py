from openpyxl import load_workbook
from random import *
wb = load_workbook("sample2.xlsx") 
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value =="영어":
            cell.value = "Python"
wb.save("sample2.xlsx")
